import sqlite3
import openpyxl

conn = sqlite3.connect("C:\\Projects\\BreezBrakBot-learn\\brakobaza_2.sqlite3")
cursor = conn.cursor()
cursor2 = conn.cursor()
async def zapolni_akt(SSCC_number, tg_id):
    cursor.execute("SELECT Doc_data, Doc_number, ER_number, SSCC_number, NS_Code, Name_item, Serial_number, Sort_select, Cell_number, Comment FROM brak WHERE SSCC_number == ?",(SSCC_number,))
    data = cursor.fetchall()
    cursor2.execute("SELECT name, town, location FROM users WHERE tg_id == ?",(tg_id,))
    data2 = cursor2.fetchall()

    workbook = openpyxl.load_workbook("C:\\Projects\\BreezBrakBot-learn\\brak_act.xlsx")
    worksheet = workbook['buffer']

    Doc_data = "A4:W6"
    worksheet.merge_cells("A4:W6")
    worksheet['A4'].value = data[0][0]

    Loc_value = "BD3:CB5"
    worksheet.merge_cells(Loc_value)
    worksheet['BD3'].value = data2[0][2]

    Town_value = "BD6:CB8"
    worksheet.merge_cells(Town_value)
    worksheet['BD6'].value = data2[0][1]

    Doc_number = "AH11:BH16"
    worksheet.merge_cells(Doc_number)
    worksheet["AH11"].value = data[0][1]

    name = "BD65:BY66"
    worksheet.merge_cells("BD65:BY66")
    worksheet["BD65"].value = data2[0][0]

    ER_value = "AV26:CB27"
    worksheet.merge_cells(ER_value)
    worksheet['AV26'].value = data[0][2]

    SSCC_value = "A33:W35"
    worksheet.merge_cells(SSCC_value)
    worksheet['A33'].value = data[0][3]

    Brak_value = "BD123:CB127"
    worksheet.merge_cells(Brak_value)
    worksheet['BD123'].value = data[0][3]

    NS_value = "A38:W40"
    worksheet.merge_cells(NS_value)
    worksheet['A38'].value = data[0][4]

    Name_value = "X33:BR40"
    worksheet.merge_cells(Name_value)
    worksheet['X33'].value = data[0][5]

    SN_value = "D44:CB46"
    worksheet.merge_cells(SN_value)
    worksheet['D44'].value = data[0][6]

    Sort_value = "A51:Q61"
    worksheet.merge_cells(Sort_value)
    worksheet['A51'].value = data[0][7]

    Comment_value = "R51:CB61"
    worksheet.merge_cells(Comment_value)
    worksheet['R51'].value = data[0][9]

    Cell_value = "B65:Z66"
    worksheet.merge_cells(Cell_value)
    worksheet['B65'].value = data[0][8]

    workbook.save("C:\\Projects\\BreezBrakBot-learn\\brak_act.xlsx")
    #conn.close()
    #conn.close()